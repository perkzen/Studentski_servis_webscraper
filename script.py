from bs4 import BeautifulSoup
import requests
from Job import Job
from openpyxl import  load_workbook

data = []


def comparePay(pay, searchParamPay, operator):
    try:
        if operator == "<":
            return searchParamPay >= float(pay.split()[0])
        elif operator == ">":
            return searchParamPay <= float(pay.split()[0])
        else:
            return True
    except AttributeError:
        return False
    except ValueError:
        return False


def compareSearchParams(compareTo, searchParam):
    if searchParam == "":
        return True
    return searchParam.strip().lower() in compareTo.strip().lower()


def writeToExcel(JobList, fileName):
    wb = load_workbook(fileName)
    ws = wb.active
    # clear old data
    ws.delete_rows(1, ws.max_row + 1)
    # insert new data
    for job in JobList:
        ws.append([job.title, job.payment, job.location, job.description, job.code])
        wb.save(fileName)


# operator (>,<) to sort pay between jobs
def fetchData(title, location, payment, operator, keywords):
    searchParams = Job(title=title, location=location, payment=payment, code=0, description=keywords)

    # need to fetch home page to get number of all pages
    urlHome = requests.get(
        'https://www.studentski-servis.com/studenti/prosta-dela?scrolltop=1&kljb=&page=1&isci=1&sort=&hourly_rate=4'
        '.98%3B26').text
    soupHome = BeautifulSoup(urlHome, 'lxml')

    # if there are a lot of pages it might take a while to scrape them all
    pages = soupHome.find_all('a', class_='page-link')[6].text

    # looping through all pages
    for page in range(int(pages)):

        url = requests.get(
            f'https://www.studentski-servis.com/studenti/prosta-dela?scrolltop=1&kljb=&page={pages}&isci=1&sort'
            f'=&hourly_rate=4 '
            '.98%3B26').text
        soup = BeautifulSoup(url, 'lxml')
        job_list = soup.find_all('article', class_='job-item')

        for job in job_list:

            title = job.find('h3').text
            payment = job.find('li', class_='job-payment').text
            description = job.find('p', class_='description').text
            location = job.find('ul', class_='job-attributes').li.text
            code = job.find('span', class_='job-code').text

            # checking if search parameters match with current job
            if compareSearchParams(title, searchParams.title) \
                    and compareSearchParams(location, searchParams.location) \
                    and compareSearchParams(description, searchParams.description) \
                    and comparePay(payment, searchParams.payment, operator):
                data.append(Job(title, payment, description, location, code))


if __name__ == '__main__':
    print("Collecting for data...")
    #   title location description pay operator keywords
    fetchData("", "", 5, ">", "")
    print("Loading Data...")
    writeToExcel(data, 'JobData.xlsx')
    print(f'{len(data)} number of jobs found!')
    print("Finished!")
